from xlsxwriter import Workbook
import json
import openpyxl


def generate_excel_file(filename, header, data):
    """:param filename : Name of the output file
       :param haeder : column names
       :param data : data
       :return
    """

    wb = Workbook(filename)
    bold = wb.add_format({'bold': True})
    ws = wb.add_worksheet("Covid Details")
    row = 0
    col = 0

    ws.write_row(0, 0, header, bold)
    row = row + 1
    for datam in data:
        ws.write_row(row, 0, datam)
        row = row + 1

    wb.close()


def read_config_file(inputfile):
    """
    reads config from file
    :return:
    """
    rows = []
    try:
        config = json.load(open(inputfile))
    except FileNotFoundError:
        print('{} file is not present.'.format(inputfile))
        return rows
    else:
        inputFileName = config['FileName']
        workbook = openpyxl.load_workbook(inputFileName)
        worksheet = workbook.active
        file_data = []
        for row in worksheet.iter_rows(values_only=True):
            (date, iso) = row
            if iso != 'iso':
                date = date.strftime('%Y-%m-%d')
                rows.append((date,iso))
        return rows
